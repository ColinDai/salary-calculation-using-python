import datetime
import openpyxl
from openpyxl.styles import Alignment
from willuse_date_judge import times_salary
from openpyxl.styles import PatternFill,Border,Side


nowTime = datetime.date.today()  # 获取今天日期，计算工龄



# 创建函数，功能是根据步骤1得到的字典、统计日期列表、example.xlsx文件路径（包含名字）和保存路径创建统计表
def get_xlsx_template(staff_name_worktimes, dates, example_xlsx_path, final_xlsx_path,dictRuYuanTime,dictGongLing,
                      dictJiBenGongzi,tuple_PaiQian):
    global row_color
    border = Border(left=Side(border_style='thin', color='000000'),
                    right=Side(border_style='thin', color='000000'),
                    top=Side(border_style='thin', color='000000'),
                    bottom=Side(border_style='thin', color='000000'))   # 创建边框样式
    row_no_fist_time = 3     # 计时器，表示从第3行开始填写数字
    example_book = openpyxl.load_workbook(example_xlsx_path)  # 读取example工作表
    example_book.save(final_xlsx_path)   # 另存至最终文件
    final_book = openpyxl.load_workbook(final_xlsx_path)  # 读取最终文件
    example_sheet = final_book['example']  # 打开example工作表
    example_sheet.column_dimensions['B'].width = 12  # 调整A列宽度
    example_sheet.column_dimensions['A'].width = 12  # 调整B列宽度
    for date1 in dates:  # 针对example工作表，在example工作表中B列，第row_no_fist_time=3行相应单元格中填入日期
        example_sheet.cell(row_no_fist_time, 2, date1)
        example_sheet.cell(row_no_fist_time, 2).alignment = Alignment(horizontal='center', vertical='center')  # 设置居中
        example_sheet.cell(row_no_fist_time, 2).border = border  # 设置边界
        example_sheet.row_dimensions[row_no_fist_time].height = 35  # 设置固定行高
        if times_salary(date1) == 3:  # 如果该日期为休息日，则加班系数为3，颜色为深绿色，设置边框
            example_sheet.cell(row_no_fist_time, 7, 3) # 幅值
            example_sheet.cell(row_no_fist_time, 7).alignment = Alignment(horizontal='center', vertical='center')  # 居中
            for row_color in range(2,8):  # 批量设置单元格颜色为草绿色，并加边框
                final_book['example'].cell(row_no_fist_time,row_color).fill = PatternFill("solid", start_color='00B050')
                final_book['example'].cell(row_no_fist_time, row_color).border=border
        elif times_salary(date1) == 2:  # 如果该日期为休息日，则加班系数为2，颜色为浅绿色，设置边框
            example_sheet.cell(row_no_fist_time, 7, 2)
            example_sheet.cell(row_no_fist_time, 7).alignment = Alignment(horizontal='center', vertical='center')
            for row_color in range(2, 8):
                final_book['example'].cell(row_no_fist_time, row_color).fill = PatternFill("solid", start_color='92D050')
                final_book['example'].cell(row_no_fist_time, row_color).border = border
        elif times_salary(date1) == 1.5:  # 如果该日期为工作日，则加班系数为1.5，不填充颜色，设置边框
            example_sheet.cell(row_no_fist_time, 7, 1.5)
            example_sheet.cell(row_no_fist_time, 7).alignment = Alignment(horizontal='center', vertical='center')
            example_sheet.cell(row_no_fist_time, 7).border = border
        final_book.save(final_xlsx_path)
        row_no_fist_time = row_no_fist_time + 1  # 日期行自加1
    final_book.save(final_xlsx_path)  # 保存文件
    for staff_name in staff_name_worktimes.keys():  # 将example工作表按统计员工复制
        final_book.copy_worksheet(final_book['example']).title = staff_name  # 复制并更改表名
        final_book[staff_name].cell(3, 1, staff_name) # 在员工名工作表中的单元格第3行 第1列 写入员工名字
        if staff_name in dictRuYuanTime.keys():  # 判断是否含有工龄数据，有则写入对应单元格
            final_book[staff_name].cell(5, 1, dictRuYuanTime[staff_name])
            final_book[staff_name].cell(7, 1, dictGongLing[staff_name])
        if staff_name in tuple_PaiQian:
            final_book[staff_name].cell(11, 1, dictGongLing[staff_name]*100)

        else:  # 没有则写入“/”
            final_book[staff_name].cell(5, 1, '/')
            final_book[staff_name].cell(7, 1, '/')
            final_book[staff_name].cell(11, 1, '/')
        final_book[staff_name].cell(3, 1).alignment = Alignment(horizontal='center', vertical='center')
        final_book[staff_name].cell(5, 1).alignment = Alignment(horizontal='center', vertical='center')
        final_book[staff_name].cell(7, 1).alignment = Alignment(horizontal='center', vertical='center')
        final_book[staff_name].cell(11, 1).alignment = Alignment(horizontal='center', vertical='center')
        if staff_name in dictJiBenGongzi.keys():
            final_book[staff_name].cell(9, 1, dictJiBenGongzi[staff_name])
        final_book[staff_name].cell(9, 1).alignment = Alignment(horizontal='center', vertical='center')
        final_book.save(final_xlsx_path)
        row_no_fist_time2 = 3  # 下面for 循环的计时器，表示填入打卡时间的开始行


        for date2 in dates:  # 在员工名工作表C列相应单元格中填入打卡日期
            da_ka_times = staff_name_worktimes[staff_name][date2]
            final_book[staff_name].cell(row_no_fist_time2, 3, da_ka_times)
            final_book[staff_name].cell(row_no_fist_time2, 3).alignment = Alignment(horizontal='centerContinuous',
                                                                                    vertical='justify',
                                                                                    wrap_text=True, shrink_to_fit=True)
            final_book[staff_name].cell(row_no_fist_time2, 3).border = border
            final_book[staff_name].row_dimensions[row_no_fist_time2].height = 35
            row_no_fist_time2 = row_no_fist_time2 + 1



    final_book.remove(final_book['example'])
    final_book.save(final_xlsx_path)
    return final_xlsx_path




