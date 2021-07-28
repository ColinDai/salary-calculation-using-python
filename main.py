import datetime
import openpyxl
import icecream as ic
import numpy as np
import chinese_calendar as ri_li
from willuse_2nd1part_get_xlsx_template import get_xlsx_template
import copy
from openpyxl.styles import PatternFill, Border, Side, Alignment
from willuse_date_judge import times_salary
from willuse_2ndpart_timesub import time_sub
from willuse_rowreturnfunction import rowreturn

nowTime = datetime.date.today()  # 获取今天日期，计算工龄


tupleNameYF = ('员工s2', '员工s4', '员工s6', '员工s8', '员工p1', '员工p3', '员工p5', '员工p7')
tuple_sann =('员工s1',  '员工s2','员工s3', '员工s4', '员工s5', '员工s6', '员工s7', '员工s8')
tuple_PQ = ('员工p1', '员工p2', '员工p3', '员工p4', '员工p5', '员工p6', '员工p7')
tuple_zj=('员工z1', '员工z2','员工z3','员工z4')
tuple_zs=('员工zs1','员工zs2','员工zs3','员工zs4','员工zs5')
# 计算工龄
# 录入开始工作的时间
dict_startwork_Time = {'员工p1': datetime.date(2009, 11, 1), '员工p2': datetime.date(2009, 10, 1),
                  '员工p3': datetime.date(2010, 5, 1), '员工p4': datetime.date(2018, 5, 1),
                  '员工p5': datetime.date(2010, 3, 1), '员工p6': datetime.date(2010, 7, 1),
                  '员工p7': datetime.date(2010, 10, 1)}
dictJiBenGongzi={'员工p1': 2220, '员工p2': 2220,
                  '员工p3': 2220, '员工p4': 2220,
                  '员工p5': 2220, '员工p6': 2220,
                  '员工p7': 2220,'员工s1':2940,
                 '员工s2':3400,'员工s3':2940, '员工s4':2940, '员工s5':2940, '员工s6':2940, '员工s7':2940}

# 计算工龄
dictGongLing = {k: nowTime.__sub__(v).days // 365 for k, v in dict_startwork_Time.items()}

# 计算时薪基数
dictJiBenGongzi_jishu2=copy.deepcopy(dictJiBenGongzi)
for staff_name in dictJiBenGongzi.keys():
    if staff_name in tuple_PQ:
        dictJiBenGongzi_jishu2[staff_name]=dictJiBenGongzi[staff_name]+dictGongLing[staff_name]*100
dictJiBenGongzi_jishu= {k: (v /21.75)/8 for k, v in dictJiBenGongzi_jishu2.items()}

"""输入1 录入需要计算的员工名字"""
staff_need_cal = ('员工s1',  '员工s2','员工s3', '员工s4', '员工s5',
                  '员工p1', '员工p2', '员工p3', '员工p4', '员工p5',
                  '员工z1', '员工z2','员工z3',
                  '员工zs1','员工zs2','员工zs3')




"""输入2  录入计算工时的起止日期"""
date_start = datetime.date(2021, 6, 22)  # 输入起始日期
date_end = datetime.date(2021, 7, 20)  # 输入截止日期
# 创建需要统计的日期列表
dates = [date_start]  # 需要计算的日期列表
while dates[-1] < date_end:
    dates.append(dates[-1] + datetime.timedelta(days=1))
# 计算统计日期中的年、月份信息
dates_months = [date_start.month]
dates_years = [date_start.year]
for date in dates:  # 计算统计日期中的年份
    dates_years.append(date.year)
dates_years = list(set(dates_years))
for date in dates:  # 计算统计日期中的月份
    dates_months.append(date.month)
dates_months = list(set(dates_months))
# 根据日期和样本表格，创建xlsx文件.......................................................................


# 根据打卡机的表格，将每个员工的打卡信息读取出来，保存在staff_name_worktime.npy这个文件中
staff_name_worktime={}
staff_worktime={}
for staff_name in staff_need_cal:
    for year in dates_years:
        for month in dates_months:
            path = 'D:\\Users\\Desktop\\2021\\工时\\' + str(year) + str(month) + '.xlsx'
            wbook = openpyxl.load_workbook(path)
            wsheet = wbook['刷卡记录']

            staff_name_row = rowreturn(wsheet)[staff_name] + 1

            for ri_qi in dates:
                if (ri_qi.month==month) and (ri_qi.year==year):
                    staff_worktime[ri_qi]=wsheet.cell(staff_name_row, ri_qi.day).value
                    # ic.ic(staff_name,year,month,path,ri_qi,staff_name_row)
    staff_name_worktime[staff_name]=copy.deepcopy(staff_worktime)
np.save('staff_name_worktime'+ str(dates_years[0]) + str(dates_months[0]) + '-' + \
                  str(dates_years[-1]) + str(dates_months[-1])+'.npy',staff_name_worktime)

# 导入计算好的字典数据（人名，日期，打卡时间）
staff_name_worktime = np.load('staff_name_worktime'+ str(dates_years[0]) + str(dates_months[0]) + '-' + \
                  str(dates_years[-1]) + str(dates_months[-1])+'.npy', allow_pickle='TRUE')

# 将part1中的ndarray数据转换为字典，备用
staff_name_worktimes = dict(np.ndenumerate(staff_name_worktime))[()]

# 根据字典，统计日期的年、月、日，需要统计的员工列表以及example.xlsx文件创建统计表
example_xlsx_path = 'D:\\Users\\Desktop\\2021\\工时\\example.xlsx'
final_xlsx_path = 'D:\\Users\\Desktop\\2021\\工时\\' + str(dates_years[0]) + str(dates_months[0]) + '-' + \
                  str(dates_years[-1]) + str(dates_months[-1]) + '.xlsx'

# 得到一个模板，里面包含了员工的打卡时间，工龄，基本工资等信息
final_xlsx_path2=get_xlsx_template(staff_name_worktimes, dates, example_xlsx_path, final_xlsx_path,
                  dict_startwork_Time,dictGongLing,dictJiBenGongzi,tuple_PQ)

# 读取填入打卡数据的模板工作簿
final_book2 = openpyxl.load_workbook(final_xlsx_path2)
final_book2.save('D:\\Users\\Desktop\\2021\\工时\\' + str(dates_years[0]) + str(dates_months[0]) + '-' + \
                 str(dates_years[-1]) + str(dates_months[-1]) + '时间分析.xlsx')
final_book3 = openpyxl.load_workbook(
    'D:\\Users\\Desktop\\2021\\工时\\' + str(dates_years[0]) + str(dates_months[0]) + '-' + \
    str(dates_years[-1]) + str(dates_months[-1]) + '时间分析.xlsx')
# 创建 第1部分得到的员工_日期_打卡时间字典副本
staff_name_worktimes2 = copy.deepcopy(staff_name_worktimes)

# 将打卡日期由字符串型转换为time型，并和日期combine,进一步转换为datetime型
for staff_name in staff_name_worktimes2.keys():
    for date in dates:
        if staff_name_worktimes2[staff_name][date]:
            da_ka_strs = staff_name_worktimes2[staff_name][date].split()
            staff_name_worktimes2[staff_name][date] = da_ka_strs
            for No in range(len(staff_name_worktimes2[staff_name][date])):
                da_ka_str = staff_name_worktimes2[staff_name][date][No]
                da_ka_str2time = copy.deepcopy(datetime.datetime.strptime(da_ka_str, '%H:%M').time())
                staff_name_worktimes2[staff_name][date][No] = datetime.datetime.combine(date, da_ka_str2time)
# ic.ic(staff_name_worktimes2)
# 现在staff_name_worktimes2的键_值_内的列表都转换成了datetime类型，staff_name_worktimes2就是我们要进行下一步处理的字典了




# 定义几个时间区间节点
time00 = datetime.time(hour=0, minute=0)
time0015 = datetime.time(hour=0, minute=15)
time05 = datetime.time(hour=5, minute=0)
time08=datetime.time(hour=8, minute=0)
time12 = datetime.time(hour=12, minute=0)
time13 = datetime.time(hour=13, minute=0)
time1645 = datetime.time(hour=16, minute=45)
time17 = datetime.time(hour=17, minute=0)
time1715 = datetime.time(hour=17, minute=15)
time19 = datetime.time(hour=19, minute=0)
time21 = datetime.time(hour=21, minute=0)
time2345 = datetime.time(hour=23, minute=45)
time2359 = datetime.time(hour=23, minute=59)


# 对taff_name_worktimes2的键_值_内的列表元素进行判断，将它给上一个日期的下班或者当天的上班或下班
# None类型不能加数据,把所有None转为[]
for staff_name in staff_name_worktimes2.keys():
    for date in dates:
        if staff_name_worktimes2[staff_name][date]==None:
            staff_name_worktimes2[staff_name][date] = []
# 把每个在00：00~05：00的时间挑出来给上一个日期的上班，并在字典副本中把这个值删掉
for staff_name in staff_name_worktimes2.keys():
    for date in dates:
        need_del = []
        if staff_name_worktimes2[staff_name][date]==None:
            staff_name_worktimes2[staff_name][date]=[]    # None类型不能加数据
        if staff_name_worktimes2[staff_name][date]:
            for da_ka_time in staff_name_worktimes2[staff_name][date]: # 对于每个员工-日期-打卡时间
                if time00 <= da_ka_time.time() <= time05:  # 如果打卡时间在00-0500
                    need_del.append(da_ka_time)  # 保存这个datetime
                    if (date - datetime.timedelta(days=1)) in dates: # 如果前一天在计算的日期内
                        staff_name_worktimes2[staff_name][(date - datetime.timedelta(days=1))].append(da_ka_time)
                        # 上一个日期的结尾加上将该次打卡datetime
                        date_row = dates.index(date) + 3 - 1
                        final_book3[staff_name].cell(date_row, 5, da_ka_time)  # 上一个日期的下班日填入该打卡时间
                    staff_name_worktimes2[staff_name][date] = \
                        list(set(staff_name_worktimes2[staff_name][date]) - set(need_del))  # 去除掉在00:00~05:00的时间
            staff_name_worktimes2[staff_name][date].sort()


# 处理打卡记录中的重复打卡值
for staff_name in staff_name_worktimes2.keys():
    for date in dates:
        rep_time=[]
        if staff_name_worktimes2[staff_name][date]: # 有值才有长度
            if len(staff_name_worktimes2[staff_name][date]) >= 2:  # 如果打卡记录值大于等于2个，则开始比较
                for i in range(len(staff_name_worktimes2[staff_name][date])):
                    for j in range(i + 1, len(staff_name_worktimes2[staff_name][date])):
                        if abs(staff_name_worktimes2[staff_name][date][i] -
                               staff_name_worktimes2[staff_name][date][j])<= datetime.timedelta(seconds=900):
                            rep_time.append(staff_name_worktimes2[staff_name][date][i])
                    staff_name_worktimes2[staff_name][date] = list(set(staff_name_worktimes2[staff_name][date]) -
                                                               set(rep_time))
            staff_name_worktimes2[staff_name][date].sort()
# 两个打卡时间的差值小于15min（900秒），把重复值中的小值去掉





final_book3.save('D:\\Users\\Desktop\\2021\\工时\\' + str(dates_years[0]) + str(
        dates_months[0]) + '-' + str(dates_years[-1]) + str(dates_months[-1]) + '时间分析.xlsx')





border = Border(left=Side(border_style='thin', color='000000'),
                right=Side(border_style='thin', color='000000'),
                top=Side(border_style='thin', color='000000'),
                bottom=Side(border_style='thin', color='000000'))  # 创建边框样式
# no_fill = openpyxl.styles.PatternFill(fill_type=None)  # 创建无格式单元格

"""输入3：输入可能上夜班的人员名单"""
dict_yeban_name=['员工p1', '员工p2', '员工s1','员工s2']  # 所有有可能上夜班的人员放在这个列表里，言外之意这个列表之外的人不可能上夜班
dict_ye_ban_date_name =copy.deepcopy(staff_name_worktimes2)  # 包含是否上夜班的字典


# 根据员工姓名和日期判断该员工该天是否上夜班
for staff_name in staff_name_worktimes2.keys():
    for date in dates:
        dict_ye_ban_date_name[staff_name][date] = False
        if staff_name in dict_yeban_name:  # 如果该员工在列表里，也即可能上夜班，则进一步判断打卡时间
            if bool(staff_name_worktimes2[staff_name][date]):  # 如果当天有打卡时间
                if time13 < staff_name_worktimes2[staff_name][date][0].time() <= time19 or \
                        time21 < staff_name_worktimes2[staff_name][date][-1].time() <= time2359 or \
                        time00 < staff_name_worktimes2[staff_name][date][-1].time() <= time05:  # 第一个值13-19或最后一个值21-05
                    dict_ye_ban_date_name[staff_name][date] = True
                if time05 < staff_name_worktimes2[staff_name][date][0].time() < time13 or \
                        time13 < staff_name_worktimes2[staff_name][date][-1].time() < time19:
                    dict_ye_ban_date_name[staff_name][date] = False
                if len(staff_name_worktimes2[staff_name][date]) == 1:
                    if time05 < staff_name_worktimes2[staff_name][date][0].time() <= time21:
                        dict_ye_ban_date_name[staff_name][date] = False
# 得到一个是否上夜班的字典
# 创建夜班函数，在打卡时间人工校正后进一步判断是否上夜班


# 对员工 日期 打卡时间进行初步的自动分析
for staff_name in staff_name_worktimes2.keys():  # 在处理过00:00-05:00的数据且构建完夜班函数后，对字典中的值进行上下班的分析
    for date in dates:
        date_row2 = dates.index(date) + 3  # 根据日期索引的行数+3，从第3行开始数
        if staff_name_worktimes2[staff_name][date] :  # 打卡记录有值

            if len(staff_name_worktimes2[staff_name][date]) == 2:  # 打卡日期有2个值的情况
                final_book3[staff_name].cell(date_row2, 4, staff_name_worktimes2[staff_name][date][0])
                final_book3[staff_name].cell(date_row2, 5, staff_name_worktimes2[staff_name][date][-1])

            elif len(staff_name_worktimes2[staff_name][date]) == 1:  # 打卡日期为1个值的情况
                if dict_ye_ban_date_name[staff_name][date]==False:  # 不上夜班
                    if time05 < staff_name_worktimes2[staff_name][date][0].time() <= time12:  # 不上夜班且在12点以前给上班
                        final_book3[staff_name].cell(date_row2, 4, staff_name_worktimes2[staff_name][date][0])
                        final_book3[staff_name].cell(date_row2, 4).border = border
                        # 下班时间没值 下班给默认下班时间17.00并标红
                        if final_book3[staff_name].cell(date_row2, 5).value is None:
                            final_book3[staff_name].cell(date_row2, 5,
                                                         datetime.datetime.combine(date,
                                                                                   datetime.time(hour=17, minute=0)))
                            final_book3[staff_name].cell(date_row2, 5).border = border
                            final_book3[staff_name].cell(date_row2, 5).fill = PatternFill("solid", start_color='FF0000')
                        elif final_book3[staff_name].cell(date_row2, 5).value:  # 下班时间有值
                            final_book3[staff_name].cell(date_row2, 5).fill = PatternFill("solid",
                                                                                          start_color='00FFFFFF')
                    if time12 <= staff_name_worktimes2[staff_name][date][0].time() <= time2359 and \
                            (final_book3[staff_name].cell(date_row2, 5).value is None):  # 在12点到23点59且第二个没值
                        final_book3[staff_name].cell(date_row2, 5, staff_name_worktimes2[staff_name][date][0])
                        final_book3[staff_name].cell(date_row2, 5).border = border
                        # 上班给默认时间8.00并标红
                        final_book3[staff_name].cell(date_row2, 4,
                                                     datetime.datetime.combine(date, datetime.time(hour=8, minute=0)))
                        final_book3[staff_name].cell(date_row2, 4).border = border
                        final_book3[staff_name].cell(date_row2, 4).fill = PatternFill("solid", start_color='FF0000')
                if dict_ye_ban_date_name[staff_name][date]==True:  # 上夜班
                    if time05 < staff_name_worktimes2[staff_name][date][0].time() <= time21:  # 该时间为上班时间
                        final_book3[staff_name].cell(date_row2, 4, staff_name_worktimes2[staff_name][date][0])
                        final_book3[staff_name].cell(date_row2, 4).border = border
                        # 上夜班  下班未打卡  但下班时间有值 将单元格设置成没颜色
                        if final_book3[staff_name].cell(date_row2, 5).value:
                            final_book3[staff_name].cell(date_row2, 5).fill = PatternFill("solid",
                                                                                          start_color='00FFFFFF')
                        # 上夜班  下班未打卡  下班时间无值 将下班时间默认写入 23：45 并标红
                        elif final_book3[staff_name].cell(date_row2, 5).value is None:
                            final_book3[staff_name].cell(date_row2, 5,
                                                         datetime.datetime.combine(date,
                                                                                   datetime.time(hour=23, minute=45)))
                            final_book3[staff_name].cell(date_row2, 5).border = border
                            final_book3[staff_name].cell(date_row2, 5).fill = PatternFill("solid",
                                                                                          start_color='00FFFFFF')
                    if (time21 <= staff_name_worktimes2[staff_name][date][0].time() <= time2359 or
                            time00 <= staff_name_worktimes2[staff_name][date][0].time() <= time05):  # 该时间为下班时间
                        final_book3[staff_name].cell(date_row2, 4,
                                                     datetime.datetime.combine(date,
                                                                               datetime.time(hour=16, minute=30)))
                        final_book3[staff_name].cell(date_row2, 4).border = border
                        final_book3[staff_name].cell(date_row2, 4).fill = PatternFill("solid",
                                                                                      start_color='FF0000')
                        final_book3[staff_name].cell(date_row2, 5, staff_name_worktimes2[staff_name][date][0])
                        final_book3[staff_name].cell(date_row2, 5).border = border

            elif len(staff_name_worktimes2[staff_name][date]) >= 3:  # 打卡日期大于等于3个情况 第一个值给上班，最后一个值给下班
                final_book3[staff_name].cell(date_row2, 4, staff_name_worktimes2[staff_name][date][0])
                if final_book3[staff_name].cell(date_row2, 5).value is None:  # 下班格子没值则把最后一个值给下班
                    final_book3[staff_name].cell(date_row2, 5, staff_name_worktimes2[staff_name][date][-1])



final_book3.save('D:\\Users\\Desktop\\2021\\工时\\' + str(dates_years[0]) + str(
    dates_months[0]) + '-' + str(dates_years[-1]) + str(dates_months[-1]) + '时间分析.xlsx')
final_book3.save('D:\\Users\\Desktop\\2021\\工时\\' + str(dates_years[0]) + str(
    dates_months[0]) + '-' + str(dates_years[-1]) + str(dates_months[-1]) + '人工校正.xlsx')

"""输入6：重命名为人工校正后"""
input_test=''
while input_test!="Yes":
    input_test=input('打开后缀为“人工校正”的工作表，校正打卡时间，另存为(将"人工校正"替换为"人工校正后")后,关闭文件输入Yes:')


"""第三部分开始，根据校正过的打卡时间，计算加班小时数"""
final_book4 = openpyxl.load_workbook(
    'D:\\Users\\Desktop\\2021\\工时\\' + str(dates_years[0]) + str(dates_months[0]) + '-' + \
    str(dates_years[-1]) + str(dates_months[-1]) + '人工校正后.xlsx')
start_row = 3  # 需要计算的起始行
find_date_row = {}  # 得到一个字典，键为日期，值是日期所在的行
for date in dates:
    find_date_row[date] = start_row
    start_row = start_row + 1  # done


staff_name_date_overwork_time = {}  # 创建一个字典，key分别为（1.5,2,3,出勤天数,缺勤天数）
for staff_name in staff_name_worktimes2.keys():
    staff_name_date_overwork_time[staff_name] = {'1.5': 0, '2': 0, '3': 0, 'att_days':0,'abs_days':0}  # done

# 创建夜班函数，在打卡时间人工校正的基础上进一步判断是否上夜班,此时上下班有且仅有一个值
def ye_ban_After_correction(shang_ban_time,xia_ban_time,staff_name,dict_yeban_name):
        if staff_name in dict_yeban_name:  # 如果该员工在列表里，也即可能上夜班，则进一步判断打卡时间
            if bool(shang_ban_time) and bool(xia_ban_time):  # 如果当天上下班均有打卡时间
                if time13 < shang_ban_time.time() <= time2359 and (time21 < xia_ban_time.time() <= time2359 or
                        time00 <= xia_ban_time.time() <= time05):
                    # 上班时间13-19或下班时间21-05
                    return True
                else:
                    return False
            else:
                return False
        else:
            return False
# 得到一个是否上夜班的字典
#



for staff_name in staff_name_worktimes2.keys():  # 开始循环 统计各种类型的加班时间
    data_absence_start_row = 8  # 出勤天数起始行在第8行
    absence_days = 0  # 表示缺勤天数的初始值
    attendance_days = 0  # 表示出勤天数的初始值
    for date in dates:
        date_row2 = dates.index(date) + 3  # 根据日期索引的行数+3，从第3行开始数
        if final_book4[staff_name].cell(find_date_row[date], 4).value and \
                final_book4[staff_name].cell(find_date_row[date], 5).value:  # 上、下班时间都有值
            attendance_days = attendance_days + 1  # 有值则出勤天数+1
            shang_ban_time=final_book4[staff_name].cell(find_date_row[date], 4).value
            xia_ban_time=final_book4[staff_name].cell(find_date_row[date], 5).value
            if ye_ban_After_correction(shang_ban_time,xia_ban_time,staff_name,dict_yeban_name)==False:  # 上白班的情况
                if times_salary(date) == 1.5:  # 如果是工作日，那么加班工资系数为1.5
                    date_15xia_ban = final_book4[staff_name].cell(find_date_row[date], 5).value
                    date_time1715 = datetime.datetime.combine(date, time1715)  # 当天下午17:15
                    date_time1645 = datetime.datetime.combine(date, time1645)  # 当天下午16:45
                    date_time0500 = datetime.datetime.combine(date + datetime.timedelta(days=1), time05)  # 第二天凌晨5:00
                    if date_time1715 <= date_15xia_ban <= date_time0500:  # 下班时间有值且下班时间大于17:15小于第二天凌晨05:00
                        # 计算加班时间
                        hours_1point5 = time_sub(date_time1645,date_15xia_ban)
                        final_book4[staff_name].cell(find_date_row[date], 6, hours_1point5)
                        final_book4[staff_name].cell(find_date_row[date], 6).border = border
                        final_book4[staff_name].cell(find_date_row[date], 6).alignment = \
                            Alignment(horizontal='center', vertical='center')
                        staff_name_date_overwork_time[staff_name]['1.5'] = \
                            staff_name_date_overwork_time[staff_name]['1.5'] + hours_1point5  # 累加计算平时加班时间


                elif times_salary(date) == 2:
                    date_2shang_ban = final_book4[staff_name].cell(find_date_row[date], 4).value
                    date_2xia_ban = final_book4[staff_name].cell(find_date_row[date], 5).value
                    hours_2 = time_sub(date_2shang_ban, date_2xia_ban)
                    final_book4[staff_name].cell(find_date_row[date], 6, hours_2)
                    final_book4[staff_name].cell(find_date_row[date], 6).border = border
                    final_book4[staff_name].cell(find_date_row[date], 6).alignment = \
                        Alignment(horizontal='center', vertical='center')
                    staff_name_date_overwork_time[staff_name]['2'] = \
                        staff_name_date_overwork_time[staff_name]['2'] + hours_2

                elif times_salary(date) == 3:
                        date_3shang_ban = final_book4[staff_name].cell(find_date_row[date], 4).value
                        date_3xia_ban = final_book4[staff_name].cell(find_date_row[date], 5).value
                        hours_3 = time_sub(date_3shang_ban, date_3xia_ban)
                        final_book4[staff_name].cell(find_date_row[date], 6, hours_3)
                        final_book4[staff_name].cell(find_date_row[date], 6).border = border
                        final_book4[staff_name].cell(find_date_row[date], 6).alignment = \
                            Alignment(horizontal='center', vertical='center')
                        staff_name_date_overwork_time[staff_name]['3'] = \
                            staff_name_date_overwork_time[staff_name]['3'] + hours_3

            elif ye_ban_After_correction(shang_ban_time,xia_ban_time,staff_name,dict_yeban_name)==True:  # 上夜班的情况
                if times_salary(date) == 1.5:  # 如果是工作日，那么加班工资系数为1.5
                    date_ye_15xia_ban = final_book4[staff_name].cell(find_date_row[date], 5).value
                    date_time0015 = datetime.datetime.combine(date + datetime.timedelta(days=1), time0015)  #
                    date_time2345 = datetime.datetime.combine(date, time2345)  # 当天下午23:45
                    # date_time05 = datetime.datetime.combine(date + datetime.timedelta(days=1), time05)  # 当天下午17:15
                    if date_time0015 <= date_ye_15xia_ban:  # 下班时间有值且下班时间大于00:15小于05
                        # 计算加班时间
                        hours_ye_1point5 =time_sub(date_time2345,date_ye_15xia_ban)
                        final_book4[staff_name].cell(find_date_row[date], 6, hours_ye_1point5)
                        final_book4[staff_name].cell(find_date_row[date], 6).border = border
                        final_book4[staff_name].cell(find_date_row[date], 6).alignment = \
                            Alignment(horizontal='center', vertical='center')
                        staff_name_date_overwork_time[staff_name]['1.5'] = \
                            staff_name_date_overwork_time[staff_name]['1.5'] + hours_ye_1point5  # 累加计算平时加班时间

                elif times_salary(date) == 2:
                    date_ye_2shang_ban = final_book4[staff_name].cell(find_date_row[date], 4).value
                    date_ye_2xia_ban = final_book4[staff_name].cell(find_date_row[date], 5).value
                    if date_ye_2xia_ban.time() >= time0015:
                        hours_ye_2 = time_sub(date_ye_2shang_ban, date_ye_2xia_ban)
                        final_book4[staff_name].cell(find_date_row[date], 6, hours_ye_2)
                        final_book4[staff_name].cell(find_date_row[date], 6).border = border
                        final_book4[staff_name].cell(find_date_row[date], 6).alignment = \
                            Alignment(horizontal='center', vertical='center')
                        staff_name_date_overwork_time[staff_name]['2'] = \
                            staff_name_date_overwork_time[staff_name]['2'] + hours_ye_2

                elif times_salary(date) == 3:
                    date_ye_3shang_ban = final_book4[staff_name].cell(find_date_row[date], 4).value
                    date_ye_3xia_ban = final_book4[staff_name].cell(find_date_row[date], 5).value
                    if date_ye_3xia_ban.time() >= time0015:
                        hours_ye_3 = time_sub(date_ye_3shang_ban, date_ye_3xia_ban)
                        final_book4[staff_name].cell(find_date_row[date], 6, hours_ye_3)
                        final_book4[staff_name].cell(find_date_row[date], 6).border = border
                        final_book4[staff_name].cell(find_date_row[date], 6).alignment = \
                            Alignment(horizontal='center', vertical='center')
                        staff_name_date_overwork_time[staff_name]['3'] = \
                            staff_name_date_overwork_time[staff_name]['3'] + hours_ye_3
        elif  final_book4[staff_name].cell(find_date_row[date], 4).value == None and \
                final_book4[staff_name].cell(find_date_row[date], 5).value == None:  # 上、下班时间都无值且为工作日
            if ri_li.is_workday(date):
                absence_days =absence_days +1
                final_book4[staff_name].cell(date_row2, 4).fill = PatternFill("solid", start_color='FF0000')
                final_book4[staff_name].cell(date_row2, 5).fill = PatternFill("solid", start_color='FF0000')
                final_book4[staff_name].cell(data_absence_start_row, 9, date)  # 写入缺勤的日期
                final_book4[staff_name].cell(data_absence_start_row, 9).alignment = \
                    Alignment(horizontal='center', vertical='center')  # 设置居中
                final_book4[staff_name].cell(data_absence_start_row, 9).border = border  # 设置边框
                data_absence_start_row = data_absence_start_row + 1  # 下一个缺勤日期的行数自加1
                final_book4[staff_name].cell(data_absence_start_row, 9).alignment = \
                    Alignment(horizontal='center', vertical='center')

    staff_name_worktimes2[staff_name]['att_days']=attendance_days  # 出勤天数信息保存在字典中
    staff_name_worktimes2[staff_name]['abs_days']=absence_days  # 缺勤天数信息保存在字典中
    final_book4[staff_name].cell(8, 8, attendance_days)  # 写入出勤天数
    final_book4[staff_name].cell(8, 8).border = border
    final_book4[staff_name].cell(8, 8).alignment = Alignment(horizontal='center', vertical='center')
    final_book4[staff_name].cell(8, 10, absence_days)  # 写入缺勤天数
    final_book4[staff_name].cell(8, 10).border = border
    final_book4[staff_name].cell(8, 10).alignment = Alignment(horizontal='center', vertical='center')

    final_book4[staff_name].cell(3, 9, staff_name_date_overwork_time[staff_name]['1.5'])  # 写入平时加班时间

    final_book4[staff_name].cell(4, 9, staff_name_date_overwork_time[staff_name]['2'])  # 写入周末加班时间

    final_book4[staff_name].cell(5, 9, staff_name_date_overwork_time[staff_name]['3'])  # 写入法定节假日加班时间
    if (staff_name in tuple_PQ) or (staff_name in tuple_sann):  # 如果员工为非正式员工
        final_book4[staff_name].cell(3, 10, round(staff_name_date_overwork_time[staff_name]['1.5']*
                                     dictJiBenGongzi_jishu[staff_name]*1.5))   # 写入平时加班工资
        final_book4[staff_name].cell(4, 10, round(staff_name_date_overwork_time[staff_name]['2'] *
                                     dictJiBenGongzi_jishu[staff_name] * 2))  # 写入周末加班工资
        final_book4[staff_name].cell(5, 10, round(staff_name_date_overwork_time[staff_name]['3']*
                                 dictJiBenGongzi_jishu[staff_name]*3))   # 写入法定节假日加班工资
    final_book4[staff_name].cell(5, 9).alignment = Alignment(horizontal='center', vertical='center')
    final_book4[staff_name].cell(5, 9).border = border
    final_book4[staff_name].cell(3, 9).alignment = Alignment(horizontal='center', vertical='center')
    final_book4[staff_name].cell(3, 9).border = border
    final_book4[staff_name].cell(4, 9).alignment = Alignment(horizontal='center', vertical='center')
    final_book4[staff_name].cell(4, 9).border = border
    final_book4[staff_name].cell(5, 10).alignment = Alignment(horizontal='center', vertical='center')
    final_book4[staff_name].cell(5, 10).border = border
    final_book4[staff_name].cell(3, 10).alignment = Alignment(horizontal='center', vertical='center')
    final_book4[staff_name].cell(3, 10).border = border
    final_book4[staff_name].cell(4, 10).alignment = Alignment(horizontal='center', vertical='center')
    final_book4[staff_name].cell(4, 10).border = border

final_book4.save('D:\\Users\\Desktop\\2021\\工时\\' + str(dates_years[0]) + str(dates_months[0]) + '-' + \
                 str(dates_years[-1]) + str(dates_months[-1]) + '人工校正后.xlsx')
final_book4.save('D:\\Users\\Desktop\\2021\\工时\\' + str(dates_years[0]) + str(dates_months[0]) + '-' + \
                 str(dates_years[-1]) + str(dates_months[-1]) + '人工校正后+.xlsx')


"""输入6：对加班工时进行进一步校正"""
input_test=''
while input_test!="Yes2":
    input_test=input('打开确认后缀为“人工校正+”的文件中的加班工时，缺勤信息等，完成后另存为"人工校正后+1"输入Yes2:')


staff_name_date_overwork_time2=copy.deepcopy(staff_name_date_overwork_time)

# 根据校正后的加班工时对总表进行填写
final_book5 = openpyxl.load_workbook('D:\\Users\\Desktop\\2021\\工时\\' + str(dates_years[0]) + str
              (dates_months[0]) + '-' + str(dates_years[-1]) + str(dates_months[-1]) + '人工校正后+1.xlsx')

# 根据第二次加班工时校正得到的加班时间和缺勤信息，对保存这些字典的信息进行改写
for staff_name in staff_name_date_overwork_time2.keys():
    staff_name_date_overwork_time2[staff_name]['1.5']=final_book5[staff_name].cell(3,9).value
    staff_name_date_overwork_time2[staff_name]['2'] = final_book5[staff_name].cell(4, 9).value
    staff_name_date_overwork_time2[staff_name]['3'] = final_book5[staff_name].cell(5, 9).value
    staff_name_date_overwork_time2[staff_name]['att_days'] = final_book5[staff_name].cell(8, 8).value  # 出勤天数信息保存在字典中
    staff_name_date_overwork_time2[staff_name]['abs_days'] = final_book5[staff_name].cell(8, 10).value
zong_biao_name_row={}
zhengshi_biao_name_row={}
final_book5_max_row=56 # 表总表的最大行数
final_book5_zhengshibiao_max_row=9  # 表正式工加班的最大行数
for row in range(5, final_book5_max_row):
    zong_biao_name_row[final_book5['总表'].cell(row, 2).value] = row  # zong_biao_name_row['姓名']=姓名所在的行

for row in range(3, final_book5_max_row):
    zhengshi_biao_name_row[final_book5['正式工加班'].cell(row, 2).value] = row  # zong_biao_name_row['正式工加班']=姓名所在的行

for staff_name in staff_name_worktimes2.keys():
    # 写入平时工资
    if staff_name in tuple_sann:  # 三年制与派遣员工表格不同，所以需要分开讨论.员工如果是三年制的，则
        final_book5['总表'].cell(zong_biao_name_row[staff_name],5,round(
            staff_name_date_overwork_time2[staff_name]['1.5']*1.5*dictJiBenGongzi_jishu[staff_name]))
        # 写入假期和周末工资
        final_book5['总表'].cell(zong_biao_name_row[staff_name], 6,round(
                               staff_name_date_overwork_time2[staff_name]['2']*2*dictJiBenGongzi_jishu[staff_name]+
                               staff_name_date_overwork_time2[staff_name]['3']*3*dictJiBenGongzi_jishu[staff_name]))
        if staff_name in tupleNameYF:
            final_book5['总表'].cell(zong_biao_name_row[staff_name], 7, round(staff_name_date_overwork_time2[staff_name]
                                                                            ['2'] +staff_name_date_overwork_time2
                                                                            [staff_name]['att_days']*10))
        if staff_name_date_overwork_time2[staff_name]['abs_days']==0:  # 缺勤天数==0，给200全勤奖
            final_book5['总表'].cell(zong_biao_name_row[staff_name], 9, 200)
        final_book5.save('D:\\Users\\Desktop\\2021\\工时\\' + str(dates_years[0]) + str
        (dates_months[0]) + '-' + str(dates_years[-1]) + str(dates_months[-1]) + 'Final.xlsx')
    elif staff_name in tuple_PQ:
        final_book5['总表'].cell(zong_biao_name_row[staff_name], 4, dictGongLing[staff_name]*100)
        final_book5['总表'].cell(zong_biao_name_row[staff_name], 5,round(
                               staff_name_date_overwork_time2[staff_name]['1.5']*1.5*dictJiBenGongzi_jishu[staff_name]))
        # 写入假期和周末工资
        final_book5['总表'].cell(zong_biao_name_row[staff_name], 6,round(
                               staff_name_date_overwork_time2[staff_name]['2']*2*dictJiBenGongzi_jishu[staff_name] +
                               staff_name_date_overwork_time2[staff_name]['3']*3*dictJiBenGongzi_jishu[staff_name]))
        if staff_name in tupleNameYF:
            final_book5['总表'].cell(zong_biao_name_row[staff_name], 7, round(
                staff_name_date_overwork_time2[staff_name]['2'] +
                staff_name_date_overwork_time2[staff_name]['att_days'] * 10))
        if staff_name_date_overwork_time2[staff_name]['abs_days'] == 0:  # 缺勤天数==0，给200全勤奖
            final_book5['总表'].cell(zong_biao_name_row[staff_name], 9, 200)
        final_book5.save('D:\\Users\\Desktop\\2021\\工时\\' + str(dates_years[0]) + str
        (dates_months[0]) + '-' + str(dates_years[-1]) + str(dates_months[-1]) + 'Final.xlsx')
    elif staff_name in tuple_zj:   # 镇江的计算工作天数和加班时间
        try:
            final_book5['总表'].cell(zong_biao_name_row[staff_name], 3,
                                   staff_name_date_overwork_time2[staff_name]['att_days'])
        except KeyError:
            print('总表中无' + staff_name + '的相关信息,如仍需计算，请修改example.xlsx再试')
        else:
            staff_zhenjiang_overwok_str = '平时' + str(staff_name_date_overwork_time2[staff_name]['1.5']) + '小时，周末' + \
                                          str(staff_name_date_overwork_time2[staff_name]['2']) + '小时，节假日' + \
                                          str(staff_name_date_overwork_time2[staff_name]['3']) + '小时'
            final_book5['总表'].cell(zong_biao_name_row[staff_name], 4, staff_zhenjiang_overwok_str)
            final_book5.save('D:\\Users\\Desktop\\2021\\工时\\' + str(dates_years[0]) + str
            (dates_months[0]) + '-' + str(dates_years[-1]) + str(dates_months[-1]) + 'Final.xlsx')
    elif staff_name in tuple_zs:  # 在“正式工加班表中”填写正式员工的加班时间
        final_book5['正式工加班'].cell(zhengshi_biao_name_row[staff_name],3,
                                  staff_name_date_overwork_time2[staff_name]['1.5'])
        final_book5['正式工加班'].cell(zhengshi_biao_name_row[staff_name], 4,
                                  staff_name_date_overwork_time2[staff_name]['2'])
        final_book5['正式工加班'].cell(zhengshi_biao_name_row[staff_name], 5,
                                  staff_name_date_overwork_time2[staff_name]['3'])
        final_book5['正式工加班'].cell(zhengshi_biao_name_row[staff_name], 6,
                                  staff_name_date_overwork_time2[staff_name]['1.5']+
                                  staff_name_date_overwork_time2[staff_name]['2']+
                                  staff_name_date_overwork_time2[staff_name]['3'])
        final_book5['正式工加班'].cell(zhengshi_biao_name_row[staff_name],3).alignment =\
                                                                    Alignment(horizontal='center', vertical='center')
        final_book5['正式工加班'].cell(zhengshi_biao_name_row[staff_name], 4).alignment = \
                                                                    Alignment(horizontal='center', vertical='center')
        final_book5['正式工加班'].cell(zhengshi_biao_name_row[staff_name], 5).alignment = \
                                                                    Alignment(horizontal='center', vertical='center')
        final_book5['正式工加班'].cell(zhengshi_biao_name_row[staff_name], 6).alignment = \
                                                                    Alignment(horizontal='center', vertical='center')
        final_book5.save('D:\\Users\\Desktop\\2021\\工时\\' + str(dates_years[0]) + str
        (dates_months[0]) + '-' + str(dates_years[-1]) + str(dates_months[-1]) + 'Final.xlsx')

final_book5.save('D:\\Users\\Desktop\\2021\\工时\\' + str(dates_years[0]) + str
              (dates_months[0]) + '-' + str(dates_years[-1]) + str(dates_months[-1]) + 'Final.xlsx')


